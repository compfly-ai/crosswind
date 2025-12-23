"""Text extraction from various file formats using Docling.

This module provides text extraction capabilities for context documents:
- PDF: Uses Docling for advanced layout analysis, table extraction, OCR
- DOCX/PPTX/XLSX: Native Docling support
- CSV/Excel: Converts to readable text format
- Markdown/Text: Direct reading
- JSON: Pretty-printed with key paths

The extracted text is stored back in MongoDB (ExtractedText field) and can be
used by the scenario generator to create targeted test scenarios.

Docling (MIT License) is used for document processing:
https://github.com/docling-project/docling
"""

import csv
import io
import json
import logging
import tempfile
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# Maximum characters to extract per file (to stay within LLM context limits)
MAX_CHARS_PER_FILE = 100_000
# Maximum total characters across all files in a context
MAX_CHARS_TOTAL = 500_000


class BaseExtractor(ABC):
    """Base class for text extractors."""

    @abstractmethod
    def extract(self, content: bytes) -> tuple[str, dict]:
        """Extract text from file content.

        Args:
            content: Raw file bytes

        Returns:
            Tuple of (extracted_text, metadata_dict)
            metadata_dict may include: page_count, row_count, char_count, etc.
        """
        pass

    def truncate_text(self, text: str, max_chars: int = MAX_CHARS_PER_FILE) -> str:
        """Truncate text to max characters with indicator."""
        if len(text) <= max_chars:
            return text
        return text[:max_chars] + f"\n\n[... truncated, {len(text) - max_chars} chars omitted ...]"


class DoclingExtractor(BaseExtractor):
    """Extract text from documents using Docling (PDF, DOCX, PPTX, images)."""

    def __init__(self):
        self._converter = None

    @property
    def converter(self):
        """Lazy-load the Docling converter."""
        if self._converter is None:
            try:
                from docling.document_converter import DocumentConverter
                self._converter = DocumentConverter()
            except ImportError:
                logger.error("Docling not installed. Install with: pip install docling")
                raise
        return self._converter

    def extract(self, content: bytes, suffix: str = ".pdf") -> tuple[str, dict]:
        """Extract text using Docling.

        Args:
            content: Raw file bytes
            suffix: File extension (e.g., ".pdf", ".docx")
        """
        try:
            # Write to temp file (Docling needs a file path)
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            try:
                # Convert document
                result = self.converter.convert(tmp_path)
                doc = result.document

                # Export to markdown (best format for LLM consumption)
                text = doc.export_to_markdown()

                # Get metadata
                metadata = {
                    "char_count": len(text),
                }

                # Try to get page count if available
                if hasattr(doc, 'pages') and doc.pages:
                    metadata["page_count"] = len(doc.pages)

                return self.truncate_text(text), metadata

            finally:
                # Clean up temp file
                Path(tmp_path).unlink(missing_ok=True)

        except Exception as e:
            logger.error(f"Docling extraction failed: {e}")
            return "", {"error": str(e)}


class CSVExtractor(BaseExtractor):
    """Extract text from CSV files."""

    def extract(self, content: bytes) -> tuple[str, dict]:
        try:
            # Try to decode as UTF-8, fall back to latin-1
            try:
                text_content = content.decode("utf-8")
            except UnicodeDecodeError:
                text_content = content.decode("latin-1")

            reader = csv.reader(io.StringIO(text_content))
            rows = list(reader)

            if not rows:
                return "", {"row_count": 0, "char_count": 0}

            # Format as readable table
            output_parts = []

            # Header row
            if rows:
                header = rows[0]
                output_parts.append("Columns: " + ", ".join(header))
                output_parts.append("")

            # Data rows (first 100 rows as sample)
            row_limit = min(100, len(rows) - 1)
            for i, row in enumerate(rows[1 : row_limit + 1], 1):
                row_text = " | ".join(
                    f"{header[j] if j < len(header) else f'col{j}'}: {cell}"
                    for j, cell in enumerate(row)
                )
                output_parts.append(f"Row {i}: {row_text}")

            if len(rows) > row_limit + 1:
                output_parts.append(f"\n[... {len(rows) - row_limit - 1} more rows ...]")

            full_text = "\n".join(output_parts)
            return self.truncate_text(full_text), {
                "row_count": len(rows) - 1,  # Exclude header
                "char_count": len(full_text),
            }

        except Exception as e:
            logger.error(f"CSV extraction failed: {e}")
            return "", {"error": str(e)}


class ExcelExtractor(BaseExtractor):
    """Extract text from Excel files using openpyxl."""

    def extract(self, content: bytes) -> tuple[str, dict]:
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return "", {"error": "openpyxl not installed"}

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            output_parts = []
            total_rows = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    continue

                output_parts.append(f"=== Sheet: {sheet_name} ===")

                # Header row
                header = [str(cell) if cell else "" for cell in rows[0]]
                output_parts.append("Columns: " + ", ".join(header))
                output_parts.append("")

                # Data rows (first 50 per sheet)
                row_limit = min(50, len(rows) - 1)
                for i, row in enumerate(rows[1 : row_limit + 1], 1):
                    cells = [str(cell) if cell else "" for cell in row]
                    row_text = " | ".join(
                        f"{header[j] if j < len(header) else f'col{j}'}: {cell}"
                        for j, cell in enumerate(cells)
                    )
                    output_parts.append(f"Row {i}: {row_text}")

                if len(rows) > row_limit + 1:
                    output_parts.append(f"\n[... {len(rows) - row_limit - 1} more rows in this sheet ...]")

                total_rows += len(rows) - 1
                output_parts.append("")

            workbook.close()

            full_text = "\n".join(output_parts)
            return self.truncate_text(full_text), {
                "row_count": total_rows,
                "char_count": len(full_text),
            }

        except Exception as e:
            logger.error(f"Excel extraction failed: {e}")
            return "", {"error": str(e)}


class MarkdownExtractor(BaseExtractor):
    """Extract text from Markdown files (direct pass-through)."""

    def extract(self, content: bytes) -> tuple[str, dict]:
        try:
            try:
                text = content.decode("utf-8")
            except UnicodeDecodeError:
                text = content.decode("latin-1")

            return self.truncate_text(text), {"char_count": len(text)}

        except Exception as e:
            logger.error(f"Markdown extraction failed: {e}")
            return "", {"error": str(e)}


class JSONExtractor(BaseExtractor):
    """Extract text from JSON files with pretty formatting."""

    def extract(self, content: bytes) -> tuple[str, dict]:
        try:
            data = json.loads(content.decode("utf-8"))

            # Pretty print with indentation
            text = json.dumps(data, indent=2, ensure_ascii=False)

            return self.truncate_text(text), {"char_count": len(text)}

        except Exception as e:
            logger.error(f"JSON extraction failed: {e}")
            return "", {"error": str(e)}


class TextExtractor:
    """Main text extractor that routes to appropriate handler."""

    # Content types that Docling handles well
    DOCLING_TYPES = {
        "application/pdf": ".pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation": ".pptx",
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/tiff": ".tiff",
    }

    # Other extractors
    SIMPLE_EXTRACTORS = {
        "text/csv": CSVExtractor,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ExcelExtractor,
        "application/vnd.ms-excel": ExcelExtractor,
        "text/markdown": MarkdownExtractor,
        "text/x-markdown": MarkdownExtractor,
        "text/plain": MarkdownExtractor,
        "application/json": JSONExtractor,
    }

    def __init__(self):
        self._docling_extractor = None

    @property
    def docling_extractor(self):
        """Lazy-load Docling extractor."""
        if self._docling_extractor is None:
            self._docling_extractor = DoclingExtractor()
        return self._docling_extractor

    def extract(self, content: bytes, content_type: str) -> tuple[str, dict]:
        """Extract text from file content based on content type.

        Args:
            content: Raw file bytes
            content_type: MIME type of the file

        Returns:
            Tuple of (extracted_text, metadata_dict)
        """
        # Check if Docling handles this type
        if content_type in self.DOCLING_TYPES:
            suffix = self.DOCLING_TYPES[content_type]
            return self.docling_extractor.extract(content, suffix=suffix)

        # Check simple extractors
        extractor_class = self.SIMPLE_EXTRACTORS.get(content_type)
        if extractor_class:
            extractor = extractor_class()
            return extractor.extract(content)

        logger.warning(f"No extractor for content type: {content_type}")
        return "", {"error": f"Unsupported content type: {content_type}"}


def extract_text_from_file(file_path: str, content_type: Optional[str] = None) -> tuple[str, dict]:
    """Convenience function to extract text from a file path.

    Args:
        file_path: Path to the file
        content_type: Optional MIME type (will be guessed from extension if not provided)

    Returns:
        Tuple of (extracted_text, metadata_dict)
    """
    path = Path(file_path)

    if not content_type:
        # Guess content type from extension
        ext_to_type = {
            ".pdf": "application/pdf",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            ".csv": "text/csv",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xls": "application/vnd.ms-excel",
            ".md": "text/markdown",
            ".txt": "text/plain",
            ".json": "application/json",
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".tiff": "image/tiff",
        }
        content_type = ext_to_type.get(path.suffix.lower(), "application/octet-stream")

    with open(path, "rb") as f:
        content = f.read()

    extractor = TextExtractor()
    return extractor.extract(content, content_type)
